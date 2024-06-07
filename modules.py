import requests
import json
import openpyxl
import os


url = os.getenv("URL")
cashier_bolt_id = os.getenv("BOLT_ID")
balti_store_id = os.getenv("BALTI_ID")
kristiine_store_id = os.getenv("KRISTINE_ID")
payment_bolt_id = os.getenv("PAYMENT_ID")
km_id = os.getenv("KM_ID")
token = os.getenv("TOKEN")
docs = os.getenv("PATH_DOCS")
table_kristiine = os.getenv("SHEET_KRISTIINE")
table_balti = os.getenv("SHEET_BALTI")
table_items = os.getenv("SHEET_ITEMS")


headers = {
    "Content-Type": "application/json",
    "Authorization": f"Bearer {token}"
}
response = requests.get(url=url, headers=headers)
items_lv = response.json()['items']


path_of_docs = docs
sheet_balti = openpyxl.open(path_of_docs + table_balti, read_only=True).active
sheet_kristiine = openpyxl.open(path_of_docs + table_kristiine, read_only=True).active
sheet_items = openpyxl.open(path_of_docs + table_items, read_only=True).active


#module for appending IDs from variants by SKU
sku = 'ET015'
id_test = 0
index=0
for i in range(len(dict['items'])+1):
    if sku == dict['items'][index]['variants'][0]['sku']:
        id_test = dict['items'][index]['variants'][0]['variant_id']
    else:
        index+=1
#module for appending IDs from variants by SKU 2
a=0
ab=0
ac=0
reciepts=[]
for i in range(len(dict['items'])+1):
        if reciepts[a]['line_items'][ac]['variant_id'] == dict['items'][ab]['variants'][0]['sku']:
            reciepts[a]['line_items'][ac]['variant_id'] = dict['items'][ab]['variants'][0]['variant_id']
            ac+=1
            ab=0
        else:
            ab+=1



#module for parsing data from excel
for row in range(1, 10):
    provider_name = sheet_items[row][0].value
    order_reference_id = sheet_items[row][4].value
    date = sheet_items[row][15].value
    sku = sheet_items[row][23].value
    amount = sheet_items[row][27].value
    #print(provider_name, order_reference_id, date, sku, amount)


#module for report weekly
receipts = []

a=[]
b=[]
for row in range(6, 10):
    if sheet_balti[row][6].value != '-':
       ab = sheet_balti[row][15].value - sheet_balti[row][11].value - sheet_balti[row][6].value
       b.append(ab)
    else:
        ab = sheet_balti[row][15].value - sheet_balti[row][11].value
        b.append(ab)

for row in range(6, 10):
    ab = sheet_balti[row][3].value
    b.append(ab)


#module for orderrefenceid
reciepts=[]
lmtms={'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}
for row in range(2, 40):
    u=len(reciepts)
    reciept={
    'store_id': None,
    'employee_id': cashier_bolt_id,
    'order': None,
    'source': None,
    'receipt_date': None,
    'line_items': [{'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}],
    'note': None,
    'payments':[{'payment_type_id': payment_bolt_id}],
    'summa': 0.0
}
    #variant1
    if sheet_items[row][4].value != sheet_items[row-1][4].value:
        reciept['order']= sheet_items[row][4].value
        reciepts.append(reciept)
        print(reciept)

    #variant2
    if sheet_items[row][4].value != None and sheet_items[row][23].value != None:
        if sheet_items[row][4].value != sheet_items[row-1][4].value:
            reciept['order']= sheet_items[row][4].value
            lmtms['variant_id'] = sheet_items[row][23].value
            reciept['line_items'] = lmtms
            reciepts.append(reciept)

    #variant3
    if sheet_items[row][4].value != None and sheet_items[row][23].value != None:
        if sheet_items[row][4].value != sheet_items[row-1][4].value:
            reciept['order']= sheet_items[row][4].value
            reciept['line_items'][0]['variant_id'] = sheet_items[row][23].value
            reciept['line_items'][0]['quantity'] = sheet_items[row][27].value
            reciept['note'] = sheet_items[row][0].value
            if sheet_items[row][0].value == 'Loco Rolls Pannkoogid Kristiine':
                reciept['store_id'] = kristiine_store_id
                reciept['source'] = 'Kristiine'
            else:
                reciept['store_id'] = balti_store_id
                reciept['source'] = 'BaltiJaam'
            if sheet_items[row][15].value != '':
                reciept['receipt_date'] = sheet_items[row][15].value
            elif sheet_items[row][15].value == '':
                reciept['receipt_date'] = sheet_items[row][9].value
            reciepts.append(reciept)
        elif sheet_items[row][4].value == sheet_items[row-1][4].value and sheet_items[row][23].value != None:
            lmtms['variant_id'] = sheet_items[row][23].value
            lmtms['quantity'] = sheet_items[row][27].value
            reciepts[u-1]['line_items'].append(lmtms)
            lmtms={'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}

# module for orders from weekly reports
for rowk in range(6, 243):
    if len(sheet_kristiine[rowk][3].value) < 7 and sheet_kristiine[rowk][3].value != '':
        print(sheet_kristiine[rowk][3].value)    


#module for % price
for i in range(len(reciepts)):
    j = len(reciepts[i]['line_items'])
    for k in range(j):
        if reciepts[i]['line_items'][k]['price']>0:
            reciepts[i]['line_items'][k]['price']=reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']
    del reciepts[i]['summa']
    
#module for summ price TEST
index_kr_wkl = 6
index_bl_wkl = 6
for i in range(len(reciepts)):
    j = len(reciepts[i]['line_items'])
    if reciepts[i]['order'] == sheet_kristiine[index_kr_wkl][3]:
        for k in range(j):
            if reciepts[i]['line_items'][k]['price']>0:
                reciepts[i]['line_items'][k]['price']=reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']
                if sheet_kristiine[index_kr_wkl+1][3] == '':
                    index_kr_wkl += 2
        del reciepts[i]['summa']
        index_kr_wkl+=1
    elif reciepts[i]['order'] == sheet_balti[index_bl_wkl][3]:
        for k in range(j):
            if reciepts[i]['line_items'][k]['price']>0:
                reciepts[i]['line_items'][k]['price']=reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']
                if sheet_balti[index_bl_wkl+1][3] == '':
                    index_bl_wkl += 2
        del reciepts[i]['summa']
        index_bl_wkl+=1

#module for summ price TEST2
index_kr_wkl = 6
index_bl_wkl = 6
for i in range(len(reciepts)):
    j = len(reciepts[i]['line_items'])
    if reciepts[i]['order'] == sheet_kristiine[index_kr_wkl][3].value:
        for k in range(j):
            if reciepts[i]['line_items'][k]['price'] > 0:
                if sheet_kristiine[index_kr_wkl][6].value != '-' and sheet_kristiine[index_kr_wkl][6].value != '':
                    kokku = sheet_kristiine[index_kr_wkl][15].value - sheet_kristiine[index_kr_wkl][11].value - sheet_kristiine[index_kr_wkl][6].value
                    reciepts[i]['line_items'][k]['price'] = (reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']) / 100
                    reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku
                if sheet_kristiine[index_kr_wkl][6].value == '-' and sheet_kristiine[index_kr_wkl][6].value != '':
                    kokku = sheet_kristiine[index_kr_wkl][15].value - sheet_kristiine[index_kr_wkl][11].value
                    reciepts[i]['line_items'][k]['price'] = (reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']) / 100
                    reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku
                if sheet_kristiine[index_kr_wkl+1][3].value == '':
                    index_kr_wkl += 2
        del reciepts[i]['summa']
        index_kr_wkl+=1
    if reciepts[i]['order'] == sheet_balti[index_bl_wkl][3].value:
        for k in range(j):
            if reciepts[i]['line_items'][k]['price'] > 0:
                if sheet_balti[index_bl_wkl][6].value != '-' and sheet_balti[index_bl_wkl][6].value != '':
                    kokku1 = sheet_balti[index_bl_wkl][15].value - sheet_balti[index_bl_wkl][11].value - sheet_balti[index_bl_wkl][6].value
                    reciepts[i]['line_items'][k]['price'] = (reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']) / 100
                    reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku1
                if sheet_balti[index_bl_wkl][6].value == '-' and sheet_balti[index_bl_wkl][6].value != '':
                    kokku1 = sheet_balti[index_bl_wkl][15].value - sheet_balti[index_bl_wkl][11].value
                    reciepts[i]['line_items'][k]['price'] = (reciepts[i]['line_items'][k]['price']*100/reciepts[i]['summa']) / 100
                    reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku1
                if sheet_balti[index_bl_wkl+1][3].value == '':
                    index_bl_wkl += 2
        del reciepts[i]['summa']
        index_bl_wkl+=1


##module for summ price FINAL
for i in range(len(reciepts)):
    j = len(reciepts[i]['line_items'])
    for index_b in range(6, sheet_balti.max_row+1):
        if reciepts[i]['order'] == sheet_balti[index_b][3].value:
            if sheet_balti[index_b][6].value != '-':
                kokku = sheet_balti[index_b][15].value - sheet_balti[index_b][11].value - sheet_balti[index_b][6].value
            else:
                kokku = sheet_balti[index_b][15].value - sheet_balti[index_b][11].value
    for index_k in range(6, sheet_kristiine.max_row+1):
        if reciepts[i]['order'] == sheet_kristiine[index_k][3].value:
            if sheet_kristiine[index_k][6].value != '-':
                kokku = sheet_kristiine[index_k][15].value - sheet_kristiine[index_k][11].value - sheet_kristiine[index_k][6].value
            else:
                kokku = sheet_kristiine[index_k][15].value - sheet_kristiine[index_k][11].value
    for k in range(j):
        if reciepts[i]['line_items'][k]['price']>0:
            reciepts[i]['line_items'][k]['price']=(reciepts[i]['line_items'][k]['price']* 100 / reciepts[i]['summa']) / 100
            reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku
    del reciepts[i]['summa']

#for i in reciepts:
  #  print(json.dumps(i, indent=4))


#TEST FINAL
reciepts=[]
lmtms={'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}
for row in range(2, sheet_items.max_row + 1):
    u=len(reciepts)
    reciept={
    'store_id': None,
    'employee_id': cashier_bolt_id,
    'order': None,
    'source': None,
    'receipt_date': None,
    'line_items': [{'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}],
    'note': None,
    'payments':[{'payment_type_id': payment_bolt_id}],
    'summa' : 0.0
}
    if sheet_items[row][4].value != None and sheet_items[row][23].value != None:
        if sheet_items[row][4].value != sheet_items[row-1][4].value:
            reciept['order']= sheet_items[row][4].value
            reciept['line_items'][0]['variant_id'] = sheet_items[row][23].value
            reciept['line_items'][0]['quantity'] = sheet_items[row][27].value
            reciept['note'] = sheet_items[row][0].value
            if sheet_items[row][34].value != None:
                reciept['line_items'][0]['price'] = sheet_items[row][34].value
                reciept['summa'] += reciept['line_items'][0]['price']
            else:
                reciept['line_items'][0]['price'] = sheet_items[row][31].value
                reciept['summa'] += reciept['line_items'][0]['price']
            if sheet_items[row][0].value == 'Loco Rolls Pannkoogid Kristiine':
                reciept['store_id'] = kristiine_store_id
                reciept['source'] = 'Kristiine'
            else:
                reciept['store_id'] = balti_store_id
                reciept['source'] = 'BaltiJaam'
            if sheet_items[row][15].value != '':
                reciept['receipt_date'] = sheet_items[row][15].value
            elif sheet_items[row][15].value == '':
                reciept['receipt_date'] = sheet_items[row][9].value
            reciepts.append(reciept)
        elif sheet_items[row][4].value == sheet_items[row-1][4].value and sheet_items[row][23].value != None:
            lmtms['variant_id'] = sheet_items[row][23].value
            lmtms['quantity'] = sheet_items[row][27].value
            if sheet_items[row][34].value != None:
                lmtms['price'] = sheet_items[row][34].value
                reciepts[u-1]['summa'] += lmtms['price']
            else:
                lmtms['price'] = sheet_items[row][31].value
                reciepts[u-1]['summa'] += lmtms['price']
            reciepts[u-1]['line_items'].append(lmtms)
        lmtms={'variant_id': None, 'quantity': None, 'price': None, 'line_taxes': [{'id':km_id}]}

#a = len(reciepts)
for i in range(len(reciepts)):
    j = len(reciepts[i]['line_items'])
    for index_b in range(6, sheet_balti.max_row + 1):
        if reciepts[i]['order'] == sheet_balti[index_b][3].value:
            if sheet_balti[index_b][6].value != '-':
                kokku = sheet_balti[index_b][15].value - sheet_balti[index_b][11].value - sheet_balti[index_b][6].value
            else:
                kokku = sheet_balti[index_b][15].value - sheet_balti[index_b][11].value
    for index_k in range(6, sheet_kristiine.max_row + 1):
        if reciepts[i]['order'] == sheet_kristiine[index_k][3].value:
            if sheet_kristiine[index_k][6].value != '-':
                kokku = sheet_kristiine[index_k][15].value - sheet_kristiine[index_k][11].value - sheet_kristiine[index_k][6].value
            else:
                kokku = sheet_kristiine[index_k][15].value - sheet_kristiine[index_k][11].value
    for k in range(j):
        if reciepts[i]['line_items'][k]['price'] > 0:
            reciepts[i]['line_items'][k]['price']=(reciepts[i]['line_items'][k]['price']* 100 / reciepts[i]['summa']) / 100
            reciepts[i]['line_items'][k]['price'] = reciepts[i]['line_items'][k]['price'] * kokku
        for req in range(len(items_lv) + 1):
            if reciepts[i]['line_items'][k]['variant_id'] == items_lv[req-1]['variants'][0]['sku']:
                reciepts[i]['line_items'][k]['variant_id'] = items_lv[req-1]['variants'][0]['variant_id']
    del reciepts[i]['summa']


#CHECKING
summa_grand = 0
len_test=len(reciepts)
for poi in range(len_test):
    len_test2=len(reciepts[poi]['line_items'])
    for iop in range(len_test2):
        if reciepts[poi]['line_items'][iop]['price'] > 0:
            if reciepts[poi]['line_items'][iop]['quantity'] == 1:
                summa_grand+=reciepts[poi]['line_items'][iop]['price']
            elif reciepts[poi]['line_items'][iop]['quantity'] > 1:
                summa_grand+=reciepts[poi]['line_items'][iop]['price'] * reciepts[poi]['line_items'][iop]['quantity']

print(round(summa_grand, 2))
print(len(reciepts))


